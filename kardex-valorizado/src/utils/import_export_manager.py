"""
Manager Centralizado para Importación y Exportación de Plantillas
Archivo: src/utils/import_export_manager.py
"""
import re
import calendar
from datetime import datetime, date
from decimal import Decimal, ROUND_HALF_UP

from PyQt6.QtWidgets import QFileDialog, QMessageBox
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from models.database_model import (obtener_session, Proveedor, Producto, Compra,
                                   Venta, Cliente, Categoria, TipoCambio,
                                   TipoDocumento, Moneda, Equipo, TipoEquipo,
                                   SubtipoEquipo, Almacen, NivelEquipo, EstadoEquipo)

# Definición de unidades SUNAT
UNIDADES_SUNAT = [
    "UND - Unidad", "KG - Kilogramo", "GR - Gramo", "LT - Litro",
    "ML - Mililitro", "M - Metro", "M2 - Metro cuadrado", "M3 - Metro cúbico",
    "CJ - Caja", "PQ - Paquete", "BOL - Bolsa", "SAC - Saco",
    "GLN - Galón", "DOC - Docena", "MIL - Millar"
]

class ImportExportManager:
    def __init__(self, parent_widget=None):
        self.parent = parent_widget
        self.session = obtener_session()

    def generar_plantilla(self, modulo):
        """Dispara el método de generación de plantilla según el módulo."""
        if modulo == "Proveedores":
            self._generar_plantilla_proveedores()
        elif modulo == "Clientes":
            self._generar_plantilla_clientes()
        elif modulo == "Productos":
            self._generar_plantilla_productos()
        elif modulo == "Equipos":
            self._generar_plantilla_equipos()
        elif modulo == "Compras":
            self._generar_plantilla_compras()
        elif modulo == "Ventas":
            self._generar_plantilla_ventas()
        elif modulo == "Tipo de Cambio":
            self._generar_plantilla_tipo_cambio()
        else:
            QMessageBox.warning(self.parent, "Error", f"El módulo '{modulo}' no es válido para generar plantillas.")

    def importar_datos(self, modulo):
        """Dispara el método de importación de datos según el módulo."""
        if modulo == "Proveedores":
            self._importar_datos_proveedores()
        elif modulo == "Clientes":
            self._importar_datos_clientes()
        elif modulo == "Productos":
            self._importar_datos_productos()
        elif modulo == "Equipos":
            self._importar_datos_equipos()
        elif modulo == "Compras":
            self._importar_datos_compras()
        elif modulo == "Ventas":
            self._importar_datos_ventas()
        elif modulo == "Tipo de Cambio":
            self._importar_datos_tipo_cambio()
        else:
            QMessageBox.warning(self.parent, "Error", f"El módulo '{modulo}' no es válido para importar datos.")

    def _generar_plantilla_proveedores(self):
        """Crea y guarda una plantilla de Excel para la importación de proveedores."""
        file_path, _ = QFileDialog.getSaveFileName(
            self.parent, "Guardar Plantilla de Proveedores",
            "plantilla_proveedores.xlsx", "Archivos de Excel (*.xlsx)"
        )

        if not file_path:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Proveedores"
            headers = ["RUC", "RAZON_SOCIAL", "DIRECCION", "TELEFONO", "EMAIL", "CONTACTO"]
            ws.append(headers)

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="FF1D6F42", end_color="FF1D6F42", fill_type="solid")
            header_align = Alignment(horizontal="center")

            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align

            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 40
            ws.column_dimensions['C'].width = 40
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 30
            ws.column_dimensions['F'].width = 30

            ws_inst = wb.create_sheet(title="Instrucciones")
            ws_inst.append(["Columna", "Descripción", "Obligatorio"])
            ws_inst.append(["RUC", "RUC de 11 dígitos. Debe empezar con 10, 15, 17 o 20.", "Sí"])
            ws_inst.append(["RAZON_SOCIAL", "Nombre o Razón Social del proveedor.", "Sí"])
            ws_inst.append(["DIRECCION", "Dirección fiscal.", "No"])
            ws_inst.append(["TELEFONO", "Teléfono de contacto.", "No"])
            ws_inst.append(["EMAIL", "Correo electrónico.", "No"])
            ws_inst.append(["CONTACTO", "Nombre de la persona de contacto.", "No"])

            wb.save(file_path)
            QMessageBox.information(self.parent, "Éxito", f"Plantilla guardada en:\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self.parent, "Error", f"No se pudo guardar la plantilla:\n{str(e)}")

    def _importar_datos_proveedores(self):
        """Importa proveedores desde un archivo Excel."""
        file_path, _ = QFileDialog.getOpenFileName(
            self.parent, "Importar Proveedores", "", "Archivos de Excel (*.xlsx)"
        )
        if not file_path:
            return

        try:
            wb = load_workbook(file_path, data_only=True)
            if "Proveedores" not in wb.sheetnames:
                QMessageBox.critical(self.parent, "Error de Hoja", "No se encontró la hoja 'Proveedores'.")
                return

            ws = wb["Proveedores"]
            expected_headers = ["RUC", "RAZON_SOCIAL", "DIRECCION", "TELEFONO", "EMAIL", "CONTACTO"]
            actual_headers = [str(cell.value).upper().strip() for cell in ws[1]]
            if actual_headers != expected_headers:
                QMessageBox.critical(self.parent, "Error de Formato", f"Encabezados incorrectos.")
                return

            proveedores_excel, errores_lectura = [], []
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if all(c is None for c in row): break
                ruc, razon_social = (str(row[0]).strip() if row[0] else ""), (str(row[1]).strip() if row[1] else "")
                if not ruc and not razon_social: continue
                if not ruc or not razon_social:
                    errores_lectura.append(f"Fila {row_idx}: RUC y Razón Social son obligatorios.")
                    continue
                if len(ruc) != 11 or not ruc.startswith(('10', '15', '17', '20')):
                    errores_lectura.append(f"Fila {row_idx}: RUC '{ruc}' inválido.")
                    continue
                email = str(row[4]).strip() if row[4] else None
                if email and not re.match(r'^[\w\.-]+@[\w\.-]+\.\w+$', email):
                    errores_lectura.append(f"Fila {row_idx}: Email '{email}' inválido.")
                    continue

                proveedores_excel.append({
                    "ruc": ruc, "razon_social": razon_social,
                    "direccion": str(row[2]).strip() if row[2] else None,
                    "telefono": str(row[3]).strip() if row[3] else None,
                    "email": email, "contacto": str(row[5]).strip() if row[5] else None,
                    "fila": row_idx
                })

            if not proveedores_excel and not errores_lectura:
                QMessageBox.warning(self.parent, "Archivo Vacío", "No se encontraron datos válidos.")
                return

            rucs_excel = {p['ruc'] for p in proveedores_excel}
            existentes_db = self.session.query(Proveedor).filter(Proveedor.ruc.in_(rucs_excel)).all()
            mapa_existentes = {p.ruc: p for p in existentes_db}
            creados, actualizados = 0, 0

            for data in proveedores_excel:
                ruc = data['ruc']
                if ruc in mapa_existentes:
                    prov = mapa_existentes[ruc]
                    prov.razon_social, prov.direccion, prov.telefono, prov.email, prov.contacto, prov.activo = \
                        data['razon_social'], data['direccion'], data['telefono'], data['email'], data['contacto'], True
                    actualizados += 1
                else:
                    prov = Proveedor(**{k: v for k, v in data.items() if k != 'fila'})
                    self.session.add(prov)
                    mapa_existentes[ruc] = prov
                    creados += 1

            self.session.commit()
            self._mostrar_reporte_importacion(creados, actualizados, errores_lectura)

        except Exception as e:
            self.session.rollback()
            QMessageBox.critical(self.parent, "Error Crítico", f"Ocurrió un error inesperado:\n{str(e)}")
        finally:
            self.session.close()

    def _generar_plantilla_clientes(self):
        """Crea y guarda una plantilla de Excel para la importación de clientes."""
        file_path, _ = QFileDialog.getSaveFileName(
            self.parent, "Guardar Plantilla de Clientes",
            "plantilla_clientes.xlsx", "Archivos de Excel (*.xlsx)"
        )

        if not file_path:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Clientes"
            headers = ["RUC_DNI", "RAZON_SOCIAL_NOMBRE", "DIRECCION", "TELEFONO", "EMAIL", "CONTACTO"]
            ws.append(headers)

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="FF1D6F42", end_color="FF1D6F42", fill_type="solid")
            header_align = Alignment(horizontal="center")

            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align

            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 40
            ws.column_dimensions['C'].width = 40
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 30
            ws.column_dimensions['F'].width = 30

            ws_inst = wb.create_sheet(title="Instrucciones")
            ws_inst.append(["Columna", "Descripción", "Obligatorio"])
            ws_inst.append(["RUC_DNI", "RUC (11 dígitos) o DNI (8 dígitos).", "Sí"])
            ws_inst.append(["RAZON_SOCIAL_NOMBRE", "Nombre o Razón Social del cliente.", "Sí"])
            ws_inst.append(["DIRECCION", "Dirección fiscal o de residencia.", "No"])
            ws_inst.append(["TELEFONO", "Teléfono de contacto.", "No"])
            ws_inst.append(["EMAIL", "Correo electrónico.", "No"])
            ws_inst.append(["CONTACTO", "Nombre de la persona de contacto.", "No"])

            wb.save(file_path)
            QMessageBox.information(self.parent, "Éxito", f"Plantilla guardada en:\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self.parent, "Error", f"No se pudo guardar la plantilla:\n{str(e)}")

    def _importar_datos_clientes(self):
        """Importa clientes desde un archivo Excel."""
        file_path, _ = QFileDialog.getOpenFileName(
            self.parent, "Importar Clientes", "", "Archivos de Excel (*.xlsx)"
        )
        if not file_path:
            return

        try:
            wb = load_workbook(file_path, data_only=True)
            if "Clientes" not in wb.sheetnames:
                QMessageBox.critical(self.parent, "Error de Hoja", "No se encontró la hoja 'Clientes'.")
                return

            ws = wb["Clientes"]
            expected_headers = ["RUC_DNI", "RAZON_SOCIAL_NOMBRE", "DIRECCION", "TELEFONO", "EMAIL", "CONTACTO"]
            actual_headers = [str(cell.value).upper().strip() for cell in ws[1]]
            if actual_headers != expected_headers:
                QMessageBox.critical(self.parent, "Error de Formato", f"Encabezados incorrectos.")
                return

            clientes_excel, errores_lectura = [], []
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if all(c is None for c in row): break
                ruc_dni, razon_social = (str(row[0]).strip() if row[0] else ""), (str(row[1]).strip() if row[1] else "")

                if not ruc_dni and not razon_social: continue
                if not ruc_dni or not razon_social:
                    errores_lectura.append(f"Fila {row_idx}: RUC/DNI y Nombre son obligatorios.")
                    continue

                # Validación de RUC o DNI
                if not (len(ruc_dni) == 11 or len(ruc_dni) == 8) or not ruc_dni.isdigit():
                    errores_lectura.append(f"Fila {row_idx}: RUC/DNI '{ruc_dni}' inválido (debe ser 8 u 11 dígitos).")
                    continue

                email = str(row[4]).strip() if row[4] else None
                if email and not re.match(r'^[\w\.-]+@[\w\.-]+\.\w+$', email):
                    errores_lectura.append(f"Fila {row_idx}: Email '{email}' inválido.")
                    continue

                clientes_excel.append({
                    "ruc_o_dni": ruc_dni, "razon_social_o_nombre": razon_social,
                    "direccion": str(row[2]).strip() if row[2] else None,
                    "telefono": str(row[3]).strip() if row[3] else None,
                    "email": email, "contacto": str(row[5]).strip() if row[5] else None,
                    "fila": row_idx
                })

            if not clientes_excel and not errores_lectura:
                QMessageBox.warning(self.parent, "Archivo Vacío", "No se encontraron datos válidos.")
                return

            rucs_excel = {c['ruc_o_dni'] for c in clientes_excel}
            existentes_db = self.session.query(Cliente).filter(Cliente.ruc_o_dni.in_(rucs_excel)).all()
            mapa_existentes = {c.ruc_o_dni: c for c in existentes_db}
            creados, actualizados = 0, 0

            for data in clientes_excel:
                ruc = data['ruc_o_dni']
                if ruc in mapa_existentes:
                    cliente = mapa_existentes[ruc]
                    cliente.razon_social_o_nombre = data['razon_social_o_nombre']
                    cliente.direccion = data['direccion']
                    cliente.telefono = data['telefono']
                    cliente.email = data['email']
                    cliente.contacto = data['contacto']
                    cliente.activo = True
                    actualizados += 1
                else:
                    cliente = Cliente(**{k: v for k, v in data.items() if k != 'fila'})
                    self.session.add(cliente)
                    mapa_existentes[ruc] = cliente
                    creados += 1

            self.session.commit()
            self._mostrar_reporte_importacion(creados, actualizados, errores_lectura)

        except Exception as e:
            self.session.rollback()
            QMessageBox.critical(self.parent, "Error Crítico", f"Ocurrió un error inesperado:\n{str(e)}")
        finally:
            self.session.close()

    def _generar_plantilla_productos(self):
        """Genera una plantilla Excel para la importación masiva de productos."""
        path, _ = QFileDialog.getSaveFileName(
            self.parent, "Guardar Plantilla de Productos",
            "plantilla_productos.xlsx", "Archivos de Excel (*.xlsx)"
        )
        if not path: return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Productos"

            headers = [
                "CODIGO_PREFIJO", "NOMBRE", "DESCRIPCION", "CATEGORIA", "UNIDAD_MEDIDA",
                "STOCK_MINIMO", "PRECIO_VENTA", "MANEJA_LOTE", "MANEJA_SERIE",
                "TIENE_VENCIMIENTO", "DIAS_VENCIMIENTO"
            ]
            ws.append(headers)

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="FF1D6F42", end_color="FF1D6F42", fill_type="solid")
            header_align = Alignment(horizontal="center")
            for cell in ws[1]:
                cell.font, cell.fill, cell.alignment = header_font, header_fill, header_align

            dv_sino = DataValidation(type="list", formula1='"SI,NO"', allow_blank=True)
            ws.add_data_validation(dv_sino)
            dv_sino.add('H2:J1000')

            unidades_codigos = [u.split(' - ')[0] for u in UNIDADES_SUNAT]
            dv_unidad = DataValidation(type="list", formula1=f'"{",".join(unidades_codigos)}"', allow_blank=False)
            ws.add_data_validation(dv_unidad)
            dv_unidad.add('E2:E1000')

            categorias = self.session.query(Categoria).filter_by(activo=True).all()
            cat_nombres = [c.nombre.replace('"', "''") for c in categorias]
            if cat_nombres:
                dv_cat = DataValidation(type="list", formula1=f'"{",".join(cat_nombres)}"', allow_blank=False)
                ws.add_data_validation(dv_cat)
                dv_cat.add('D2:D1000')

            ws.column_dimensions['A'].width = 18
            ws.column_dimensions['B'].width = 40
            ws.column_dimensions['C'].width = 40
            ws.column_dimensions['D'].width = 25
            ws.column_dimensions['E'].width = 18
            for col in ['F', 'G', 'H', 'I', 'J', 'K']:
                ws.column_dimensions[col].width = 18

            ws_inst = wb.create_sheet(title="Instrucciones")
            ws_inst.append(["Columna", "Descripción", "Obligatorio"])
            ws_inst.append(["CODIGO_PREFIJO", "Prefijo de 5 caracteres para el código (Ej: TUBO0).", "Sí"])
            ws_inst.append(["NOMBRE", "Nombre del producto.", "Sí"])
            ws_inst.append(["DESCRIPCION", "Descripción detallada (opcional).", "No"])
            ws_inst.append(["CATEGORIA", "Nombre exacto de una categoría ya existente.", "Sí"])
            ws_inst.append(["UNIDAD_MEDIDA", "Código de unidad de medida (Ej: UND, KG, M).", "Sí"])
            ws_inst.append(["STOCK_MINIMO", "Valor numérico (opcional, defecto 0).", "No"])
            ws_inst.append(["PRECIO_VENTA", "Valor numérico (opcional, defecto 0).", "No"])
            ws_inst.append(["MANEJA_LOTE", "Escribir 'SI' o 'NO' (opcional, defecto NO).", "No"])
            ws_inst.append(["MANEJA_SERIE", "Escribir 'SI' o 'NO' (opcional, defecto NO).", "No"])
            ws_inst.append(["TIENE_VENCIMIENTO", "Escribir 'SI' o 'NO' (opcional, defecto NO).", "No"])
            ws_inst.append(["DIAS_VENCIMIENTO", "Número de días (solo si la columna anterior es 'SI').", "No"])

            wb.save(path)
            QMessageBox.information(self.parent, "Éxito", f"Plantilla guardada exitosamente en:\n{path}")
        except Exception as e:
            QMessageBox.critical(self.parent, "Error", f"No se pudo generar la plantilla:\n{str(e)}")

    def _importar_datos_productos(self):
        """Importa productos desde un archivo Excel."""
        path, _ = QFileDialog.getOpenFileName(
            self.parent, "Abrir Plantilla de Productos", "", "Archivos de Excel (*.xlsx *.xls)"
        )
        if not path: return

        try:
            wb = load_workbook(path, data_only=True)
            if "Productos" not in wb.sheetnames:
                QMessageBox.critical(self.parent, "Error de Hoja", "No se encontró la hoja 'Productos'.")
                return

            ws = wb["Productos"]
            expected_headers = [
                "CODIGO_PREFIJO", "NOMBRE", "DESCRIPCION", "CATEGORIA", "UNIDAD_MEDIDA",
                "STOCK_MINIMO", "PRECIO_VENTA", "MANEJA_LOTE", "MANEJA_SERIE",
                "TIENE_VENCIMIENTO", "DIAS_VENCIMIENTO"
            ]
            actual_headers = [str(cell.value).upper().strip() for cell in ws[1]][:len(expected_headers)]
            if actual_headers != expected_headers:
                QMessageBox.critical(self.parent, "Error de Formato", "Los encabezados no son correctos.")
                return

            productos_a_crear, errores_lectura = [], []
            categorias_db = self.session.query(Categoria).filter_by(activo=True).all()
            cat_map = {cat.nombre.upper(): cat.id for cat in categorias_db}
            unidades_validas = [u.split(' - ')[0] for u in UNIDADES_SUNAT]
            nombres_prefijos_excel = set()

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if all(c is None for c in row): break
                try:
                    codigo_prefijo = str(row[0]).strip().upper() if row[0] else ""
                    nombre = str(row[1]).strip() if row[1] else ""
                    if not codigo_prefijo and not nombre: continue
                    if len(codigo_prefijo) != 5: raise ValueError(f"Código prefijo debe tener 5 caracteres.")
                    if not nombre: raise ValueError("El Nombre es obligatorio.")

                    categoria_nombre = str(row[3]).strip().upper() if row[3] else ""
                    categoria_id = cat_map.get(categoria_nombre)
                    if not categoria_id: raise ValueError(f"Categoría '{row[3]}' no encontrada.")

                    unidad = str(row[4]).strip().upper() if row[4] else ""
                    if unidad not in unidades_validas: raise ValueError(f"Unidad_Medida '{unidad}' no válida.")

                    stock_min = float(str(row[5]).strip() if row[5] else "0")
                    precio_venta = float(str(row[6]).strip() if row[6] else "0")

                    tiene_lote = (str(row[7]).strip().upper() if row[7] else "NO") == "SI"
                    tiene_serie = (str(row[8]).strip().upper() if row[8] else "NO") == "SI"
                    tiene_vencimiento = (str(row[9]).strip().upper() if row[9] else "NO") == "SI"

                    dias_vencimiento = int(str(row[10]).strip() if row[10] else "0") if tiene_vencimiento else None

                    if (codigo_prefijo, nombre) in nombres_prefijos_excel:
                        raise ValueError(f"Duplicado en Excel: Prefijo '{codigo_prefijo}' y Nombre '{nombre}'.")
                    nombres_prefijos_excel.add((codigo_prefijo, nombre))

                    productos_a_crear.append({
                        "prefijo": codigo_prefijo, "nombre": nombre, "desc": str(row[2]).strip() if row[2] else None,
                        "cat_id": categoria_id, "unidad": unidad, "stock_min": stock_min,
                        "precio": precio_venta, "lote": tiene_lote, "serie": tiene_serie,
                        "dias_venc": dias_vencimiento, "fila": row_idx
                    })
                except Exception as e:
                    errores_lectura.append(f"Fila {row_idx}: {str(e)}")

            if not productos_a_crear and not errores_lectura:
                QMessageBox.warning(self.parent, "Archivo Vacío", "No se encontraron datos válidos.")
                return

            creados = 0
            correlativos_map = {}
            nombres_prefijos_db = {(p.codigo.split('-')[0], p.nombre) for p in self.session.query(Producto).filter_by(activo=True).all()}

            for data in productos_a_crear:
                try:
                    prefijo, nombre = data['prefijo'], data['nombre']
                    if (prefijo, nombre) in nombres_prefijos_db:
                        raise ValueError(f"Ya existe en BD: Prefijo '{prefijo}', Nombre '{nombre}'.")

                    if prefijo not in correlativos_map:
                        ultimo = self.session.query(Producto).filter(Producto.codigo.like(f"{prefijo}-%")).order_by(Producto.codigo.desc()).first()
                        correlativos_map[prefijo] = int(ultimo.codigo.split('-')[1]) if ultimo else 0

                    correlativos_map[prefijo] += 1
                    codigo_completo = f"{prefijo}-{correlativos_map[prefijo]:06d}"

                    producto = Producto(
                        codigo=codigo_completo, nombre=data['nombre'], descripcion=data['desc'],
                        categoria_id=data['cat_id'], unidad_medida=data['unidad'], stock_minimo=data['stock_min'],
                        precio_venta=data['precio'] if data['precio'] > 0 else None, tiene_lote=data['lote'],
                        tiene_serie=data['serie'], dias_vencimiento=data['dias_venc']
                    )
                    self.session.add(producto)
                    creados += 1
                except Exception as e:
                    errores_lectura.append(f"Fila {data['fila']} (Procesando): {str(e)}")

            self.session.commit()
            self._mostrar_reporte_importacion(creados, 0, errores_lectura)

        except Exception as e:
            self.session.rollback()
            QMessageBox.critical(self.parent, "Error Crítico", f"Ocurrió un error inesperado:\n{str(e)}")
        finally:
            self.session.close()

    def _generar_plantilla_ventas(self):
        """Genera una plantilla Excel para la importación masiva de cabeceras de venta."""
        path, _ = QFileDialog.getSaveFileName(
            self.parent, "Guardar Plantilla de Ventas",
            "plantilla_ventas.xlsx", "Archivos de Excel (*.xlsx)"
        )
        if not path: return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Ventas"
            headers = ["NUMERO_PROCESO_CORRELATIVO", "RUC_CLIENTE", "FECHA_EMISION (dd/mm/aaaa)", "FECHA_CONTABLE (dd/mm/aaaa)", "TIPO_DOC", "SERIE", "NUMERO"]
            ws.append(headers)

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="FF1D6F42", end_color="FF1D6F42", fill_type="solid")
            for cell in ws[1]:
                cell.font, cell.fill = header_font, header_fill

            clientes = self.session.query(Cliente).filter_by(activo=True).all()
            if clientes and len(clientes) < 200:
                rucs = [c.ruc_o_dni for c in clientes]
                dv_ruc = DataValidation(type="list", formula1=f'"{",".join(rucs)}"', allow_blank=False)
                ws.add_data_validation(dv_ruc)
                dv_ruc.add('B2:B1000')

            tipos_doc_venta = ["FACTURA", "BOLETA", "NOTA_VENTA"]
            dv_tipo = DataValidation(type="list", formula1=f'"{",".join(tipos_doc_venta)}"', allow_blank=False)
            ws.add_data_validation(dv_tipo)
            dv_tipo.add('E2:E1000')

            ws_inst = wb.create_sheet(title="Instrucciones")
            ws_inst.append(["Columna", "Descripción", "Obligatorio"])
            ws_inst.append(["NUMERO_PROCESO_CORRELATIVO", "Correlativo numérico. El sistema generará el formato completo 05MMNNNNNN.", "Sí"])
            ws_inst.append(["RUC_CLIENTE", "RUC o DNI del cliente. Debe existir en la base de datos.", "Sí"])
            ws_inst.append(["FECHA_EMISION (dd/mm/aaaa)", "Fecha de la venta.", "Sí"])
            ws_inst.append(["FECHA_CONTABLE (dd/mm/aaaa)", "Fecha del periodo contable.", "Sí"])
            ws_inst.append(["TIPO_DOC", "Tipo de documento (FACTURA, BOLETA, NOTA_VENTA).", "Sí"])
            ws_inst.append(["SERIE", "Serie del documento (Ej: F001, B001, NV01).", "Sí"])
            ws_inst.append(["NUMERO", "Número del documento (Ej: 1234).", "Sí"])

            wb.save(path)
            QMessageBox.information(self.parent, "Éxito", f"Plantilla guardada en:\n{path}")
        except Exception as e:
            QMessageBox.critical(self.parent, "Error", f"No se pudo generar la plantilla:\n{str(e)}")

    def _importar_datos_ventas(self):
        """Importa cabeceras de venta desde un archivo Excel."""
        path, _ = QFileDialog.getOpenFileName(
            self.parent, "Abrir Plantilla de Ventas", "", "Archivos de Excel (*.xlsx *.xls)"
        )
        if not path: return

        try:
            wb = load_workbook(path, data_only=True)
            if "Ventas" not in wb.sheetnames:
                raise ValueError("No se encontró la hoja 'Ventas'.")

            ws = wb["Ventas"]
            expected_headers = ["NUMERO_PROCESO_CORRELATIVO", "RUC_CLIENTE", "FECHA_EMISION (DD/MM/AAAA)", "FECHA_CONTABLE (DD/MM/AAAA)", "TIPO_DOC", "SERIE", "NUMERO"]
            actual_headers = [str(cell.value).upper().strip() for cell in ws[1] if cell.value is not None]
            if not all(eh.split(' (')[0] in [ah.split(' (')[0] for ah in actual_headers] for eh in expected_headers):
                 raise ValueError("Los encabezados del Excel no son correctos.")

            ventas_a_crear, errores = [], []
            clientes_db = {c.ruc_o_dni: c.id for c in self.session.query(Cliente).filter_by(activo=True).all()}
            documentos_db = {(v.cliente_id, v.serie_documento, v.correlativo_documento) for v in self.session.query(Venta.cliente_id, Venta.serie_documento, Venta.correlativo_documento).all()}

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if all(c is None for c in row): continue
                try:
                    correlativo, ruc, fecha_emision, fecha_contable, tipo_doc, serie, numero = row[:7]
                    if not all([correlativo, ruc, fecha_emision, fecha_contable, tipo_doc, serie, numero]):
                        raise ValueError("Faltan datos obligatorios.")

                    cliente_id = clientes_db.get(str(ruc).strip())
                    if not cliente_id: raise ValueError(f"Cliente con RUC/DNI '{ruc}' no encontrado.")

                    fecha_e = (fecha_emision if isinstance(fecha_emision, datetime) else datetime.strptime(str(fecha_emision).split(" ")[0], "%d/%m/%Y")).date()
                    fecha_c = (fecha_contable if isinstance(fecha_contable, datetime) else datetime.strptime(str(fecha_contable).split(" ")[0], "%d/%m/%Y")).date()

                    n_proceso = f"05{fecha_c.month:02d}{int(correlativo):06d}"
                    serie_doc, num_doc = str(serie).strip().upper(), str(numero).strip().zfill(8)

                    if (cliente_id, serie_doc, num_doc) in documentos_db:
                        raise ValueError(f"Documento {serie_doc}-{num_doc} para este cliente ya existe.")

                    ventas_a_crear.append(Venta(
                        numero_proceso=n_proceso, cliente_id=cliente_id, fecha=fecha_e, fecha_registro_contable=fecha_c,
                        tipo_documento=TipoDocumento[tipo_doc.strip().upper()], serie_documento=serie_doc, correlativo_documento=num_doc,
                        moneda=Moneda.SOLES, tipo_cambio=Decimal('1.0'), incluye_igv=False,
                        igv_porcentaje=Decimal('18.0'), subtotal=Decimal('0.0'), igv=Decimal('0.0'), total=Decimal('0.0')
                    ))
                except Exception as e:
                    errores.append(f"Fila {row_idx}: {str(e)}")

            if errores:
                self._mostrar_reporte_importacion(0, 0, errores)
            elif ventas_a_crear:
                self.session.add_all(ventas_a_crear)
                self.session.commit()
                self._mostrar_reporte_importacion(len(ventas_a_crear), 0, [])
            else:
                QMessageBox.warning(self.parent, "Archivo Vacío", "No se encontraron datos válidos.")

        except Exception as e:
            self.session.rollback()
            QMessageBox.critical(self.parent, "Error Crítico", f"Ocurrió un error:\n{str(e)}")
        finally:
            self.session.close()

    def _generar_plantilla_compras(self):
        """Genera una plantilla Excel para la importación masiva de cabeceras de compra."""
        path, _ = QFileDialog.getSaveFileName(
            self.parent, "Guardar Plantilla de Compras",
            "plantilla_compras.xlsx", "Archivos de Excel (*.xlsx)"
        )
        if not path: return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Compras"
            headers = ["NUMERO_PROCESO_CORRELATIVO", "RUC_PROVEEDOR", "FECHA_EMISION (dd/mm/aaaa)", "FECHA_CONTABLE (dd/mm/aaaa)", "SERIE", "NUMERO"]
            ws.append(headers)

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="FF1D6F42", end_color="FF1D6F42", fill_type="solid")
            header_align = Alignment(horizontal="center")
            for cell in ws[1]:
                cell.font, cell.fill, cell.alignment = header_font, header_fill, header_align

            proveedores = self.session.query(Proveedor).filter_by(activo=True).all()
            if proveedores and len(proveedores) < 200:
                rucs = [p.ruc for p in proveedores]
                dv_ruc = DataValidation(type="list", formula1=f'"{",".join(rucs)}"', allow_blank=False)
                ws.add_data_validation(dv_ruc)
                dv_ruc.add('B2:B1000')

            ws.column_dimensions['A'].width = 28
            ws.column_dimensions['B'].width = 18
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 28
            ws.column_dimensions['E'].width = 10
            ws.column_dimensions['F'].width = 12


            ws_inst = wb.create_sheet(title="Instrucciones")
            ws_inst.append(["Columna", "Descripción", "Obligatorio"])
            ws_inst.append(["NUMERO_PROCESO_CORRELATIVO", "Correlativo numérico. El sistema generará el formato completo.", "Sí"])
            ws_inst.append(["RUC_PROVEEDOR", "RUC de 11 dígitos. Debe existir en su base de datos.", "Sí"])
            ws_inst.append(["FECHA_EMISION (dd/mm/aaaa)", "Fecha de la factura (para el Kardex).", "Sí"])
            ws_inst.append(["FECHA_CONTABLE (dd/mm/aaaa)", "Fecha del periodo contable (para el reporte).", "Sí"])
            ws_inst.append(["SERIE", "Serie de la factura (Ej: F001).", "Sí"])
            ws_inst.append(["NUMERO", "Número de la factura (Ej: 1234).", "Sí"])
            ws.append(["1", "12345678901", "30/09/2025", "01/10/2025", "F001", "1234"])

            wb.save(path)
            QMessageBox.information(self.parent, "Éxito", f"Plantilla guardada exitosamente en:\n{path}")
        except Exception as e:
            QMessageBox.critical(self.parent, "Error", f"No se pudo generar la plantilla:\n{str(e)}")

    def _importar_datos_compras(self):
        """Importa cabeceras de compra desde un archivo Excel."""
        path, _ = QFileDialog.getOpenFileName(
            self.parent, "Abrir Plantilla de Compras", "", "Archivos de Excel (*.xlsx *.xls)"
        )
        if not path: return

        try:
            wb = load_workbook(path, data_only=True)
            if "Compras" not in wb.sheetnames:
                QMessageBox.critical(self.parent, "Error de Hoja", "No se encontró la hoja 'Compras'.")
                return

            ws = wb["Compras"]
            expected_headers = ["NUMERO_PROCESO_CORRELATIVO", "RUC_PROVEEDOR", "FECHA_EMISION (DD/MM/AAAA)", "FECHA_CONTABLE (DD/MM/AAAA)", "SERIE", "NUMERO"]
            actual_headers = [str(cell.value).upper().strip() for cell in ws[1] if cell.value is not None][:len(expected_headers)]
            expected_headers_check = [h.split(' (')[0] for h in expected_headers]
            actual_headers_check = [h.split(' (')[0] for h in actual_headers]

            if actual_headers_check != expected_headers_check:
                QMessageBox.critical(self.parent, "Error de Formato", f"Los encabezados del Excel no son correctos.")
                return

            compras_a_crear, errores_lectura = [], []
            proveedores_db = self.session.query(Proveedor).filter_by(activo=True).all()
            prov_map = {p.ruc: p.id for p in proveedores_db}
            documentos_excel = set()
            documentos_db = {(c.proveedor_id, c.numero_documento) for c in self.session.query(Compra.proveedor_id, Compra.numero_documento).all()}

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if all(c is None for c in row): continue
                try:
                    correlativo_excel, ruc_excel, fecha_emision_excel, fecha_contable_excel, serie_excel, numero_excel = row[:6]
                    if not all([correlativo_excel, ruc_excel, fecha_emision_excel, fecha_contable_excel, serie_excel, numero_excel]):
                        raise ValueError("Faltan datos (Correlativo, RUC, Fechas, Serie o Número).")

                    correlativo = int(str(correlativo_excel).strip())
                    ruc = str(ruc_excel).strip()
                    if ruc not in prov_map: raise ValueError(f"Proveedor con RUC {ruc} no encontrado.")
                    proveedor_id = prov_map[ruc]

                    fecha_emision = (fecha_emision_excel if isinstance(fecha_emision_excel, datetime) else datetime.strptime(str(fecha_emision_excel).split(" ")[0], "%d/%m/%Y")).date()
                    fecha_contable = (fecha_contable_excel if isinstance(fecha_contable_excel, datetime) else datetime.strptime(str(fecha_contable_excel).split(" ")[0], "%d/%m/%Y")).date()

                    mes_contable = fecha_contable.month
                    numero_proceso_completo = f"06{mes_contable:02d}{correlativo:06d}"

                    serie = str(serie_excel).strip().upper()
                    numero = str(numero_excel).strip().zfill(8)
                    numero_documento_completo = f"{serie}-{numero}"
                    doc_key = (proveedor_id, numero_documento_completo)
                    if doc_key in documentos_excel: raise ValueError(f"Documento {numero_documento_completo} duplicado en el archivo.")
                    documentos_excel.add(doc_key)
                    if doc_key in documentos_db: raise ValueError(f"Documento {numero_documento_completo} ya existe en la BD.")

                    compras_a_crear.append(Compra(
                        numero_proceso=numero_proceso_completo,
                        proveedor_id=proveedor_id, fecha=fecha_emision, fecha_registro_contable=fecha_contable,
                        tipo_documento=TipoDocumento.FACTURA, numero_documento=numero_documento_completo,
                        moneda=Moneda.SOLES, tipo_cambio=Decimal('1.0'), incluye_igv=False,
                        igv_porcentaje=Decimal('18.0'), subtotal=Decimal('0.0'), igv=Decimal('0.0'), total=Decimal('0.0')
                    ))
                except Exception as e:
                    errores_lectura.append(f"Fila {row_idx}: {str(e)}")

            if errores_lectura:
                self.session.rollback()
                self._mostrar_reporte_importacion(0, 0, errores_lectura)
            elif not compras_a_crear:
                QMessageBox.warning(self.parent, "Archivo Vacío", "No se encontraron datos válidos.")
            else:
                self.session.add_all(compras_a_crear)
                self.session.commit()
                self._mostrar_reporte_importacion(len(compras_a_crear), 0, [])
        except Exception as e:
            self.session.rollback()
            QMessageBox.critical(self.parent, "Error Crítico", f"Ocurrió un error inesperado:\n{str(e)}")
        finally:
            self.session.close()

    def _generar_plantilla_equipos(self):
        """Genera una plantilla Excel para la importación masiva de equipos."""
        path, _ = QFileDialog.getSaveFileName(
            self.parent, "Guardar Plantilla de Equipos",
            "plantilla_equipos.xlsx", "Archivos de Excel (*.xlsx)"
        )
        if not path: return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Equipos"

            headers = [
                "CODIGO_PREFIJO", "NOMBRE", "TIPO_EQUIPO", "SUBTIPO_EQUIPO", "MARCA",
                "MODELO", "SERIE", "CAPACIDAD", "NIVEL", "ALMACEN", "ESTADO",
                "PROVEEDOR_RUC", "TARIFA_SOLES", "TARIFA_DOLARES", "DESCRIPCION"
            ]
            ws.append(headers)

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="FF1D6F42", end_color="FF1D6F42", fill_type="solid")
            header_align = Alignment(horizontal="center")
            for cell in ws[1]:
                cell.font, cell.fill, cell.alignment = header_font, header_fill, header_align

            # Validaciones de datos

            # Niveles
            niveles = [n.value for n in NivelEquipo]
            dv_nivel = DataValidation(type="list", formula1=f'"{",".join(niveles)}"', allow_blank=False)
            ws.add_data_validation(dv_nivel)
            dv_nivel.add('I2:I1000')

            # Estados
            estados = [e.value for e in EstadoEquipo]
            dv_estado = DataValidation(type="list", formula1=f'"{",".join(estados)}"', allow_blank=False)
            ws.add_data_validation(dv_estado)
            dv_estado.add('K2:K1000')

            # Almacenes
            almacenes = self.session.query(Almacen).filter_by(activo=True).all()
            alm_nombres = [a.nombre.replace('"', "''") for a in almacenes]
            if alm_nombres:
                dv_alm = DataValidation(type="list", formula1=f'"{",".join(alm_nombres)}"', allow_blank=False)
                ws.add_data_validation(dv_alm)
                dv_alm.add('J2:J1000')

            # Tipos de Equipo
            tipos = self.session.query(TipoEquipo).filter_by(activo=True).all()
            tipo_nombres = [t.nombre.replace('"', "''") for t in tipos]
            if tipo_nombres:
                dv_tipo = DataValidation(type="list", formula1=f'"{",".join(tipo_nombres)}"', allow_blank=False)
                ws.add_data_validation(dv_tipo)
                dv_tipo.add('C2:C1000')

            ws.column_dimensions['A'].width = 18
            ws.column_dimensions['B'].width = 40
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 25
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 20
            ws.column_dimensions['G'].width = 20
            ws.column_dimensions['H'].width = 15
            ws.column_dimensions['I'].width = 15
            ws.column_dimensions['J'].width = 25
            ws.column_dimensions['K'].width = 15
            ws.column_dimensions['L'].width = 15
            ws.column_dimensions['M'].width = 15
            ws.column_dimensions['N'].width = 15
            ws.column_dimensions['O'].width = 40

            ws_inst = wb.create_sheet(title="Instrucciones")
            ws_inst.append(["Columna", "Descripción", "Obligatorio"])
            ws_inst.append(["CODIGO_PREFIJO", "Prefijo de 5 caracteres para el código (Ej: GENER).", "Sí"])
            ws_inst.append(["NOMBRE", "Nombre del equipo.", "Sí"])
            ws_inst.append(["TIPO_EQUIPO", "Nombre exacto del tipo de equipo.", "Sí"])
            ws_inst.append(["SUBTIPO_EQUIPO", "Nombre exacto del subtipo (opcional).", "No"])
            ws_inst.append(["MARCA", "Marca del equipo.", "No"])
            ws_inst.append(["MODELO", "Modelo del equipo.", "No"])
            ws_inst.append(["SERIE", "Número de serie.", "No"])
            ws_inst.append(["CAPACIDAD", "Capacidad (Ej: 5000W).", "No"])
            ws_inst.append(["NIVEL", "Nivel de jerarquía (NIVEL_A, NIVEL_B, etc).", "Sí"])
            ws_inst.append(["ALMACEN", "Nombre del almacén donde se ubica.", "No"])
            ws_inst.append(["ESTADO", "Estado actual (DISPONIBLE, ALQUILADO, etc).", "Sí"])
            ws_inst.append(["PROVEEDOR_RUC", "RUC del proveedor/propietario.", "No"])
            ws_inst.append(["TARIFA_SOLES", "Tarifa diaria en Soles.", "No"])
            ws_inst.append(["TARIFA_DOLARES", "Tarifa diaria en Dólares.", "No"])
            ws_inst.append(["DESCRIPCION", "Descripción detallada.", "No"])

            wb.save(path)
            QMessageBox.information(self.parent, "Éxito", f"Plantilla guardada exitosamente en:\n{path}")
        except Exception as e:
            QMessageBox.critical(self.parent, "Error", f"No se pudo generar la plantilla:\n{str(e)}")

    def _importar_datos_equipos(self):
        """Importa equipos desde un archivo Excel."""
        path, _ = QFileDialog.getOpenFileName(
            self.parent, "Abrir Plantilla de Equipos", "", "Archivos de Excel (*.xlsx *.xls)"
        )
        if not path: return

        try:
            wb = load_workbook(path, data_only=True)
            if "Equipos" not in wb.sheetnames:
                QMessageBox.critical(self.parent, "Error de Hoja", "No se encontró la hoja 'Equipos'.")
                return

            ws = wb["Equipos"]
            expected_headers = [
                "CODIGO_PREFIJO", "NOMBRE", "TIPO_EQUIPO", "SUBTIPO_EQUIPO", "MARCA",
                "MODELO", "SERIE", "CAPACIDAD", "NIVEL", "ALMACEN", "ESTADO",
                "PROVEEDOR_RUC", "TARIFA_SOLES", "TARIFA_DOLARES", "DESCRIPCION"
            ]
            actual_headers = [str(cell.value).upper().strip() for cell in ws[1]][:len(expected_headers)]
            if actual_headers != expected_headers:
                QMessageBox.critical(self.parent, "Error de Formato", "Los encabezados no son correctos.")
                return

            equipos_a_crear, errores_lectura = [], []

            # Caches para búsquedas rápidas
            tipos_db = {t.nombre.upper(): t.id for t in self.session.query(TipoEquipo).filter_by(activo=True).all()}
            # Subtipos requieren búsqueda compuesta por TipoID + Nombre, se hará en el loop o precargada si no son muchos
            almacenes_db = {a.nombre.upper(): a.id for a in self.session.query(Almacen).filter_by(activo=True).all()}
            proveedores_db = {p.ruc: p.id for p in self.session.query(Proveedor).filter_by(activo=True).all()}

            # Nombres y Prefijos para validación de duplicados en Excel
            nombres_prefijos_excel = set()

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if all(c is None for c in row): break
                try:
                    # Extracción de datos
                    prefijo = str(row[0]).strip().upper() if row[0] else ""
                    nombre = str(row[1]).strip() if row[1] else ""
                    tipo_nom = str(row[2]).strip().upper() if row[2] else ""
                    subtipo_nom = str(row[3]).strip().upper() if row[3] else ""
                    marca = str(row[4]).strip() if row[4] else None
                    modelo = str(row[5]).strip() if row[5] else None
                    serie = str(row[6]).strip() if row[6] else None
                    capacidad = str(row[7]).strip() if row[7] else None
                    nivel_str = str(row[8]).strip().upper() if row[8] else ""
                    almacen_nom = str(row[9]).strip().upper() if row[9] else ""
                    estado_str = str(row[10]).strip().upper() if row[10] else "DISPONIBLE"
                    prov_ruc = str(row[11]).strip() if row[11] else None
                    tarifa_sol = float(str(row[12]).strip() if row[12] else "0")
                    tarifa_dol = float(str(row[13]).strip() if row[13] else "0")
                    desc = str(row[14]).strip() if row[14] else None

                    # Validaciones
                    if not prefijo or len(prefijo) != 5: raise ValueError("Código prefijo debe tener 5 caracteres.")
                    if not nombre: raise ValueError("Nombre es obligatorio.")
                    if not tipo_nom: raise ValueError("Tipo de Equipo es obligatorio.")
                    if not nivel_str: raise ValueError("Nivel es obligatorio.")

                    tipo_id = tipos_db.get(tipo_nom)
                    if not tipo_id: raise ValueError(f"Tipo de Equipo '{tipo_nom}' no encontrado.")

                    subtipo_id = None
                    if subtipo_nom:
                        # Buscar subtipo en BD para este tipo
                        sub = self.session.query(SubtipoEquipo).filter_by(
                            activo=True, tipo_equipo_id=tipo_id, nombre=subtipo_nom
                        ).first()
                        if sub: subtipo_id = sub.id
                        else: raise ValueError(f"Subtipo '{subtipo_nom}' no pertenece al Tipo '{tipo_nom}' o no existe.")

                    almacen_id = None
                    if almacen_nom:
                        almacen_id = almacenes_db.get(almacen_nom)
                        if not almacen_id: raise ValueError(f"Almacén '{almacen_nom}' no encontrado.")

                    prov_id = None
                    if prov_ruc:
                        prov_id = proveedores_db.get(prov_ruc)
                        if not prov_id: raise ValueError(f"Proveedor con RUC '{prov_ruc}' no encontrado.")

                    # Validar Enum Nivel y Estado
                    try:
                        nivel_enum = NivelEquipo(nivel_str)
                    except ValueError:
                        raise ValueError(f"Nivel '{nivel_str}' inválido. Use: {', '.join([n.value for n in NivelEquipo])}")

                    try:
                        estado_enum = EstadoEquipo(estado_str)
                    except ValueError:
                        raise ValueError(f"Estado '{estado_str}' inválido.")

                    if (prefijo, nombre) in nombres_prefijos_excel:
                        raise ValueError(f"Duplicado en Excel: Prefijo '{prefijo}' y Nombre '{nombre}'.")
                    nombres_prefijos_excel.add((prefijo, nombre))

                    equipos_a_crear.append({
                        "prefijo": prefijo, "nombre": nombre, "tipo_id": tipo_id, "subtipo_id": subtipo_id,
                        "marca": marca, "modelo": modelo, "serie": serie, "capacidad": capacidad,
                        "nivel": nivel_enum, "almacen_id": almacen_id, "estado": estado_enum,
                        "proveedor_id": prov_id, "tarifa_sol": tarifa_sol, "tarifa_dol": tarifa_dol,
                        "desc": desc, "fila": row_idx
                    })

                except Exception as e:
                    errores_lectura.append(f"Fila {row_idx}: {str(e)}")

            if not equipos_a_crear and not errores_lectura:
                QMessageBox.warning(self.parent, "Archivo Vacío", "No se encontraron datos válidos.")
                return

            creados = 0
            correlativos_map = {}
            # Mapa de códigos existentes para evitar duplicados de nombre/prefijo si fuera necesario
            # Pero la regla es autogenerar código.

            # Función auxiliar local para generar código sin llamar a BD cada vez (optimización)
            # Pero cuidado con concurrencia. Asumimos usuario único.

            for data in equipos_a_crear:
                try:
                    prefijo = data['prefijo']

                    if prefijo not in correlativos_map:
                        ultimo = self.session.query(Equipo).filter(Equipo.codigo.like(f"{prefijo}-%")).order_by(Equipo.codigo.desc()).first()
                        correlativos_map[prefijo] = int(ultimo.codigo.split('-')[1]) if ultimo else 0

                    correlativos_map[prefijo] += 1
                    codigo_completo = f"{prefijo}-{correlativos_map[prefijo]:06d}"

                    equipo = Equipo(
                        codigo=codigo_completo, nombre=data['nombre'], descripcion=data['desc'],
                        tipo_equipo_id=data['tipo_id'], subtipo_equipo_id=data['subtipo_id'],
                        capacidad=data['capacidad'], nivel=data['nivel'], almacen_id=data['almacen_id'],
                        estado=data['estado'], marca=data['marca'], modelo=data['modelo'],
                        serie=data['serie'], tarifa_diaria_referencial=data['tarifa_sol'],
                        tarifa_diaria_dolares=data['tarifa_dol'], proveedor_id=data['proveedor_id']
                    )
                    self.session.add(equipo)
                    creados += 1
                except Exception as e:
                    errores_lectura.append(f"Fila {data['fila']} (Procesando): {str(e)}")

            self.session.commit()
            self._mostrar_reporte_importacion(creados, 0, errores_lectura)

        except Exception as e:
            self.session.rollback()
            QMessageBox.critical(self.parent, "Error Crítico", f"Ocurrió un error inesperado:\n{str(e)}")
        finally:
            self.session.close()

    def _generar_plantilla_tipo_cambio(self):
        """Genera una plantilla Excel para la importación de tipo de cambio."""
        archivo, _ = QFileDialog.getSaveFileName(
            self.parent, "Guardar Plantilla", "plantilla_tipo_cambio.xlsx", "Excel (*.xlsx)"
        )
        if not archivo: return
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "TipoCambio"
            headers = ['fecha (yyyy-mm-dd)', 'precio_compra', 'precio_venta']
            ws.append(headers)
            ws.append(['2024-01-15', 3.850, 3.870])
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="FF1D6F42", end_color="FF1D6F42", fill_type="solid")
            wb.save(archivo)
            QMessageBox.information(self.parent, "Éxito", f"Plantilla descargada:\n{archivo}")
        except Exception as e:
            QMessageBox.critical(self.parent, "Error", f"Error al generar plantilla:\n{str(e)}")

    def _importar_datos_tipo_cambio(self):
        """Importa tipos de cambio desde Excel."""
        archivo, _ = QFileDialog.getOpenFileName(
            self.parent, "Seleccionar archivo Excel", "", "Excel (*.xlsx *.xls)"
        )
        if not archivo: return

        try:
            wb = load_workbook(archivo, data_only=True)
            ws = wb.active

            nuevos, actualizados, errores = 0, 0, []

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    if not row[0] or not row[1] or not row[2]: continue

                    fecha = (row[0].date() if isinstance(row[0], datetime) else
                             row[0] if isinstance(row[0], date) else
                             datetime.strptime(str(row[0]), '%Y-%m-%d').date())

                    precio_compra = float(row[1])
                    precio_venta = float(row[2])

                    tc_existe = self.session.query(TipoCambio).filter_by(fecha=fecha).first()
                    if tc_existe:
                        tc_existe.precio_compra, tc_existe.precio_venta, tc_existe.activo = precio_compra, precio_venta, True
                        actualizados += 1
                    else:
                        self.session.add(TipoCambio(fecha=fecha, precio_compra=precio_compra, precio_venta=precio_venta))
                        nuevos += 1
                except Exception as e:
                    errores.append(f"Fila {row_idx}: {str(e)}")

            self.session.commit()
            self._mostrar_reporte_importacion(nuevos, actualizados, errores)
        except Exception as e:
            self.session.rollback()
            QMessageBox.critical(self.parent, "Error", f"Error al importar:\n{str(e)}")
        finally:
            self.session.close()

    def _mostrar_reporte_importacion(self, creados, actualizados, errores):
        """Muestra un resumen de la operación de importación."""
        mensaje = f"Importación completada.\n\n" \
                  f"✅ Nuevos registros creados: {creados}\n" \
                  f"🔄 Registros existentes actualizados: {actualizados}\n"

        if errores:
            mensaje += f"\n⚠️ Se encontraron {len(errores)} errores que fueron omitidos:\n"
            mensaje += "\n".join(errores[:10])
            if len(errores) > 10:
                mensaje += f"\n... y {len(errores) - 10} más."
            QMessageBox.warning(self.parent, "Importación con Errores", mensaje)
        else:
            QMessageBox.information(self.parent, "Importación Exitosa", mensaje)
